"""
Export générique d'un DataFrame vers un fichier Excel formaté
(en-têtes, montants, dates, ligne de total optionnelle, largeur des
colonnes). Utilisé par tous les types d'extraction Getly : chaque module
d'extraction passe ses propres libellés de colonnes et ses colonnes
"montant" / "date" / "à totaliser".
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EN_TETE_FOND = "1F4E78"
EN_TETE_POLICE = Font(color="FFFFFF", bold=True)
TOTAL_FOND = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
TOTAL_POLICE = Font(bold=True)


def build_excel(
    df: pd.DataFrame,
    *,
    column_labels: dict[str, str] | None = None,
    montant_cols: set[str] | None = None,
    date_cols: set[str] | None = None,
    total_cols: set[str] | None = None,
    sheet_name: str = "Extraction",
) -> BytesIO:
    """
    Retourne un buffer BytesIO contenant le classeur Excel prêt à télécharger.

    - column_labels : correspondance nom technique -> libellé affiché
      (ex. {"NO_COMPTE": "N° compte"}). Les colonnes non listées gardent
      leur nom technique.
    - montant_cols : colonnes (noms techniques) formatées en nombre "0.00".
    - date_cols : colonnes (noms techniques) formatées en date jj/mm/aaaa.
    - total_cols : sous-ensemble de montant_cols pour lesquelles une ligne
      de total (somme) est ajoutée en bas du tableau.
    - sheet_name : nom de l'onglet (max 31 caractères, contrainte Excel).
    """
    column_labels = column_labels or {}
    montant_cols = montant_cols or set()
    date_cols = date_cols or set()
    total_cols = total_cols or set()
    sheet_name = (sheet_name or "Extraction")[:31]

    montant_labels = {column_labels.get(c, c) for c in montant_cols}
    date_labels = {column_labels.get(c, c) for c in date_cols}
    total_labels = {column_labels.get(c, c) for c in total_cols}

    df_export = df.rename(columns=column_labels)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        n_rows, n_cols = df_export.shape

        # --- En-têtes ---
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = EN_TETE_POLICE
            cell.fill = PatternFill(
                start_color=EN_TETE_FOND, end_color=EN_TETE_FOND, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"

        # --- Formats des colonnes montants / dates + largeur ---
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name in montant_labels:
                for row_idx in range(2, n_rows + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"
            elif col_name in date_labels:
                for row_idx in range(2, n_rows + 2):
                    ws.cell(row=row_idx, column=col_idx).number_format = "dd/mm/yyyy"

            max_len = max(
                [len(str(col_name))]
                + [len(str(v)) for v in df_export[col_name].astype(str).head(200)]
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # --- Ligne de total (somme des colonnes listées dans total_cols) ---
        if n_rows > 0 and total_labels:
            total_row = n_rows + 2
            ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_POLICE

            for col_idx, col_name in enumerate(df_export.columns, start=1):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = TOTAL_FOND
                cell.font = TOTAL_POLICE
                if col_name in total_labels:
                    col_letter = get_column_letter(col_idx)
                    cell.value = f"=SUM({col_letter}2:{col_letter}{n_rows + 1})"
                    cell.number_format = "#,##0.00"

    buffer.seek(0)
    return buffer
